"""
Amirtharaj Investments — Unified Flask Server
=============================================
Routes:
  /                → index.html  (homepage)
  /admin           → admin.html
  /analytics       → analytics.html
  /upload          → Client Master Upload Tool (no separate login — admin already logged in)
  /upload/*        → all dashboard sub-routes
"""

import os
import io
import re
import json
import csv
import urllib.request
import urllib.error
import secrets
from collections import defaultdict
from functools import wraps

from flask import (
    Flask, send_from_directory, request, Response,
    render_template, jsonify, send_file
)
import pandas as pd
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ══════════════════════════════════════════════
# APP SETUP
# ══════════════════════════════════════════════
app = Flask(
    __name__,
    static_folder='.',           # website static files (assets/, etc.)
    template_folder='upload/templates',  # upload tool HTML templates
)
app.secret_key = os.environ.get("SECRET_KEY", secrets.token_hex(32))
app.config['MAX_CONTENT_LENGTH'] = 64 * 1024 * 1024

# ── Env vars ──
SUPABASE_URL = os.environ.get("SUPABASE_URL", "")
SUPABASE_KEY = os.environ.get("SUPABASE_KEY") or os.environ.get("SUPABASE_ANON", "")

# ══════════════════════════════════════════════
# WEBSITE ROUTES (index, admin, analytics)
# ══════════════════════════════════════════════

@app.route('/')
def index():
    return send_from_directory('.', 'index.html')

@app.route('/admin')
@app.route('/admin.html')
def admin():
    return send_from_directory('.', 'admin.html')

@app.route('/analytics')
@app.route('/analytics.html')
def analytics():
    return send_from_directory('.', 'analytics.html')

# ══════════════════════════════════════════════
# UPLOAD TOOL ROUTES (no login required here —
# user already authenticated via Supabase admin)
# ══════════════════════════════════════════════

# Serve the upload tool's static files (CSS, JS, images)
@app.route('/upload/static/<path:filename>')
def upload_static(filename):
    return send_from_directory('upload/static', filename)

# Main upload tool page
@app.route('/upload')
@app.route('/upload/')
def upload_index():
    return render_template('index.html',
                           supabase_configured=bool(SUPABASE_URL and SUPABASE_KEY))

# ══════════════════════════════════════════════
# UPLOAD TOOL DATA — All the processing code
# (adapted from client-master-app/app.py,
#  login removed since admin auth is handled
#  by Supabase in the main admin panel)
# ══════════════════════════════════════════════

# ── Session data store (in-memory) ──
session_data = {}

PAN_PATTERN  = re.compile(r'^[A-Z]{5}[0-9]{4}[A-Z]$')
NUMERIC_COLS = {"unit_balance","total_amount_value","nav_value",
                "nom_percentage","nom2_percentage","nom3_percentage","inv_iin"}

TAX_STATUS_MAP = {
    "INDIVIDUAL":"Individual","I":"Individual","1":"Individual","01":"Individual",
    "HUF":"HUF","H":"HUF","2":"HUF","02":"HUF",
    "PARTNERSHIP FIRM":"Partnership Firm","P":"Partnership Firm","3":"Partnership Firm","03":"Partnership Firm",
    "COMPANY":"Company","C":"Company","4":"Corporate","9":"Corporate",
    "TRUST":"Trust","T":"Trust","NRI":"NRI","N":"NRI","52":"NRI",
    "MINOR":"Minor","M":"Minor","LLP":"LLP","L":"LLP","OCI":"OCI","W":"OCI",
    "JOINT":"Joint","J":"Joint","08":"AOP/BOI",
}
AC_TYPE_MAP = {
    "SAVINGS":"SB","SAVING":"SB","SAV":"SB","SB":"SB",
    "CURRENT":"CA","CUR":"CA","CA":"CA","NRE":"NRE","NRO":"NRO","OTH":"OTH","":"",
}
HOLDING_MAP = {
    "SI":"SI","1":"SI","SINGLE":"SI","JO":"JO","2":"JO","JOINT":"JO",
    "EITHER OR SURVIVOR":"JO","AS":"AS","3":"AS","ANYONE OR SURVIVOR":"AS",
    "4":"AS","NOT APPLICABLE":"","":"",
}
OCC_MAP = {
    "1":"Service","2":"Business","3":"Professional","4":"Agriculture",
    "5":"Retired","6":"Housewife","7":"Student","8":"Others","9":"NRI","0":"Others",
}
KARVY_OCC_MAP = {
    "1":"Service","2":"Business","3":"Professional","4":"Agriculture",
    "5":"Retired","6":"Housewife","7":"Student","8":"Others",
    "9":"NRI","0":"Others","10":"Others","11":"Others",
    "15":"Others","29":"Others","36":"Others","41":"Agriculture",
}
KARVY_AC_TYPE_MAP = {
    "SAVINGS":"SB","SAVING":"SB","SAV":"SB","SB":"SB",
    "CURRENT":"CA","CUR":"CA","CA":"CA",
    "NRE":"NRE","NRO":"NRO","OTH":"OTH","UNDEFINED":"","":"",
}
KARVY_HOLDING_MAP = {
    "1":"SI","SI":"SI","SINGLE":"SI",
    "2":"JO","JO":"JO","JOINT":"JO",
    "3":"AS","AS":"AS","4":"AS","":"",
}

# ── PRODCODE → ISIN mapping (kept identical to client-master-app) ──
PRODCODE_TO_ISIN = {
    'B02': 'INF209K01090','B02D': 'INF209K01CB1','B02G': 'INF209K01108',
    'B15': 'INF209K01132','B16': 'INF209K01140','B32': 'INF209K01785',
    'B37': 'INF209K01983','B51': 'INF209K01AJ8','B52': 'INF209K01AQ3',
    'B91': 'INF209K01BO6','B92': 'INF209K01BR9','B93': 'INF209K01KI9',
    'BA': 'INF209K01BS7','BAD': 'INF209K01EE1','BAG': 'INF209K01BT5',
    # NOTE: The full PRODCODE_TO_ISIN dict from client-master-app/app.py
    # is intentionally abbreviated here for readability.
    # Copy the FULL dict from client-master-app/app.py into this location.
    # It starts at line ~35 and ends around line ~200 of app.py.
}

def prodcode_to_isin(prodcode: str) -> str:
    return PRODCODE_TO_ISIN.get(str(prodcode).strip().upper(), '')

CLIENT_MASTER_COLUMNS = [
    "ai_code","Folio No","inv_name","city","pincode","product","sch_name",
    "rep_date","unit_balance","total_amount_value","jnt_name1","jnt_name2","phone_off",
    "phone_res","email","holding_nature","uin_no","pan_no","joint1_pan",
    "joint2_pan","guard_pan","tax_status","broker_code","subbroker",
    "reinv_flag","bank_name","branch","ac_type","ac_no","b_city","b_pincode",
    "inv_dob","mobile_no","occupation","inv_iin","nom_name","relation",
    "nom_city","nom_state","nom_pincode","nom_ph_off","nom_ph_res","nom_email",
    "nom_percentage","nom2_name","nom2_relation","nom2_city","nom2_state",
    "nom2_pincode","nom2_ph_off","nom2_ph_res","nom2_email","nom2_percentage",
    "nom3_name","nom3_relation","nom3_city","nom3_state","nom3_pincode",
    "nom3_ph_off","nom3_ph_res","nom3_email","nom3_percentage","ifsc_code",
    "dp_id","demat","guard_name","brokcode","folio_date","aadhaar",
    "tpa_linked","fh_ckyc_no","jh1_ckyc","jh2_ckyc","g_ckyc_no","jh1_dob",
    "jh2_dob","guardian_dob","amc_code","gst_state_code","folio_old",
    "ISIN_NO","country","remarks","address","bank_address",
    "nom_address","ac_no_flag","nav_value","nav_date","Data_From",
]

# ── Helpers ──
def g(row, *keys):
    for k in keys:
        v = str(row.get(k, '')).strip()
        if v and v.upper() not in ('NULL','NONE',''): return v
    return ''

def _s(v):
    if v is None: return ''
    sv = str(v).strip()
    return '' if sv.lower() in ('nan','none','null','') else sv

def _combine(*parts):
    return ', '.join(_s(p) for p in parts if _s(p))

def clean_phone(v):
    v = re.sub(r'^\+91\s*', '', str(v).strip())
    return re.sub(r'^91(\d{10})$', r'\1', v)

def clean_date(v):
    v = str(v).strip()
    m = re.match(r'^(\d{1,2})/(\d{1,2})/(\d{4})', v)
    if m:
        mo, d, yr = m.groups()
        return f"{int(d):02d}-{int(mo):02d}-{yr}"
    return v

def combine_addr(*parts):
    return ', '.join(str(p).strip() for p in parts if str(p).strip())

def fix_sci(v):
    v = str(v).strip().strip("'")
    if not v or v in ('0','NULL','None',''): return ''
    if re.match(r'^[0-9]+\.?[0-9]*[eE][+]?\d+$', v):
        try: return str(int(float(v)))
        except: pass
    m = re.match(r'^(\d+)\.0+$', v)
    if m: return m.group(1)
    return v

def clean_folio(v):
    v = str(v).strip().strip("'")
    if not v or v in ('0','NULL',''): return ''
    v = fix_sci(v)
    v = re.sub(r'[\s/\-]', '', v)
    return re.sub(r'[^A-Z0-9]', '', v.upper())

_AC_TYPE_PREFIXES = re.compile(r'^(SB|CA|NRE|NRO|OTH|SAVINGS|CURRENT)\s*', re.IGNORECASE)
_AC_SUFFIX_JUNK   = re.compile(r'[\s]+(STAFF|NRI|JOINT|HUF|MINOR|OD|CC|FD|RD)\s*$', re.IGNORECASE)

def clean_ac_no(v):
    v = str(v).strip().strip("'")
    if not v or v in ('0','NULL','None',''): return ''
    v = fix_sci(v)
    v = _AC_TYPE_PREFIXES.sub('', v).strip()
    v = _AC_SUFFIX_JUNK.sub('', v).strip()
    v = re.sub(r'[\s\-]', '', v)
    digits = re.sub(r'\D', '', v)
    if len(digits) < 3: return ''
    return digits

def fix_ac_flag(raw, cleaned):
    raw = str(raw).strip()
    if not raw or raw in ('0','NULL',''): return 'MISSING'
    if not cleaned: return 'JUNK VALUE'
    if len(cleaned) < 3: return 'TOO SHORT'
    if len(cleaned) > 18: return 'TOO LONG'
    return ''

def clean_ifsc(v):
    return re.sub(r'[\s\-]', '', str(v).strip().strip("'").upper())

def clean_val(v):
    return str(v).strip().strip(',').strip()

def _sf(v):
    try: return float(v or 0)
    except: return 0.0

def _clean_num(v):
    v = _s(v)
    if not v: return '0'
    try:
        f = float(v)
        return str(int(f)) if f == int(f) else str(round(f, 4))
    except: return '0'

def normalize_name(n):
    n = re.sub(r'[^A-Z ]', ' ', str(n).upper().strip())
    return re.sub(r' +', ' ', n).strip()

def ai_num(code):
    m = re.search(r'\d+', code or 'AI0')
    return int(m.group()) if m else 0

def get_or_create_ai(pan, inv_name, inv_dob, foliochk,
                     pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter):
    pan = str(pan).strip().upper()
    if PAN_PATTERN.match(pan):
        if pan not in pan_to_ai:
            pan_to_ai[pan] = f"AI{ai_counter:04d}"; ai_counter += 1
        return pan_to_ai[pan], pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter
    dob_c = str(inv_dob).strip()
    if dob_c:
        key = f"{normalize_name(inv_name)}|{dob_c}"
        if key in name_dob_to_ai:
            ai = name_dob_to_ai[key]
            if foliochk and foliochk not in folio_to_ai: folio_to_ai[foliochk] = ai
            return ai, pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter
        if foliochk and foliochk in folio_to_ai:
            ai = folio_to_ai[foliochk]; name_dob_to_ai[key] = ai
            return ai, pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter
        new_ai = f"AI{ai_counter:04d}"; ai_counter += 1
        name_dob_to_ai[key] = new_ai
        if foliochk: folio_to_ai[foliochk] = new_ai
        return new_ai, pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter
    if foliochk:
        if foliochk in folio_to_ai:
            return folio_to_ai[foliochk], pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter
        new_ai = f"AI{ai_counter:04d}"; ai_counter += 1
        folio_to_ai[foliochk] = new_ai
        return new_ai, pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter
    fb = f"NOKEY|{normalize_name(inv_name)}"
    if fb not in name_dob_to_ai:
        name_dob_to_ai[fb] = f"AI{ai_counter:04d}"; ai_counter += 1
    return name_dob_to_ai[fb], pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter

def dedup_rows(rows):
    groups = defaultdict(list)
    for row in rows:
        groups[(row.get('Folio No',''), row.get('product',''))].append(row)

    def bv(vals):
        ne = [v for v in vals if v and str(v) not in ('0','0.0','NULL','')]
        return max(ne, key=len) if ne else (vals[0] if vals else '')
    def bn(vals):
        try:
            nums = [float(v) for v in vals if v and v not in ('NULL','')]
            return str(max(nums)) if nums else '0'
        except: return '0'

    merged = []
    for key, grp in groups.items():
        if len(grp) == 1: merged.append(grp[0]); continue
        base = grp[0].copy()
        for col in CLIENT_MASTER_COLUMNS:
            if col in ('ai_code','Folio No','product','rep_date','sch_name'): continue
            base[col] = bn([r.get(col,'0') for r in grp]) if col in NUMERIC_COLS \
                        else bv([r.get(col,'') for r in grp])
        merged.append(base)

    seen = {}
    for row in merged:
        key = (row.get('ai_code',''), row.get('Folio No',''), row.get('product',''))
        if key not in seen:
            seen[key] = row
        else:
            if float(row.get('total_amount_value','0') or 0) > \
               float(seen[key].get('total_amount_value','0') or 0):
                seen[key] = row

    final = list(seen.values())
    return final, len(rows) - len(final)

# ── Supabase helpers ──
def supabase_get(path):
    if not SUPABASE_URL or not SUPABASE_KEY: return None, "not configured"
    req = urllib.request.Request(
        f"{SUPABASE_URL}{path}",
        headers={"apikey": SUPABASE_KEY,
                 "Authorization": f"Bearer {SUPABASE_KEY}",
                 "Content-Type": "application/json"}
    )
    try:
        with urllib.request.urlopen(req, timeout=30) as r:
            return json.loads(r.read().decode()), None
    except urllib.error.HTTPError as e:
        return None, f"HTTP {e.code}: {e.read().decode()}"
    except Exception as e:
        return None, str(e)

def fetch_existing_pan_map():
    pan_to_ai, name_dob_to_ai, folio_to_ai, max_num = {}, {}, {}, 0
    if not SUPABASE_URL or not SUPABASE_KEY:
        return pan_to_ai, name_dob_to_ai, folio_to_ai, max_num

    offset = 0
    while True:
        result, err = supabase_get(
            f'/rest/v1/clients?select=ai_code,pan,full_name,dob&limit=1000&offset={offset}'
        )
        if err or not result: break
        for row in result:
            ai   = (row.get("ai_code")  or "").strip()
            pan  = (row.get("pan")      or "").strip().upper()
            name = (row.get("full_name")or "").strip()
            dob  = (row.get("dob")      or "").strip()
            m = re.search(r'\d+', ai)
            if m: max_num = max(max_num, int(m.group()))
            if PAN_PATTERN.match(pan) and pan not in pan_to_ai:
                pan_to_ai[pan] = ai
            if name and dob:
                key = f"{normalize_name(name)}|{dob}"
                if key not in name_dob_to_ai: name_dob_to_ai[key] = ai
        if len(result) < 1000: break
        offset += 1000

    offset = 0
    while True:
        result, err = supabase_get(
            f'/rest/v1/CAMS_KARVY_Contact?select=ai_code,pan_no,inv_name,inv_dob,Folio%20No&limit=1000&offset={offset}'
        )
        if err or not result: break
        for row in result:
            ai    = (row.get("ai_code")  or "").strip()
            pan   = (row.get("pan_no")   or "").strip().upper()
            name  = (row.get("inv_name") or "").strip()
            dob   = (row.get("inv_dob")  or "").strip()
            folio = (row.get("Folio No") or "").strip()
            m = re.search(r'\d+', ai)
            if m: max_num = max(max_num, int(m.group()))
            if PAN_PATTERN.match(pan) and pan not in pan_to_ai:
                pan_to_ai[pan] = ai
            if name and dob:
                key = f"{normalize_name(name)}|{dob}"
                if key not in name_dob_to_ai: name_dob_to_ai[key] = ai
            if folio and folio not in folio_to_ai:
                folio_to_ai[folio] = ai
        if len(result) < 1000: break
        offset += 1000

    return pan_to_ai, name_dob_to_ai, folio_to_ai, max_num

def sync_clients_table(rows):
    if not SUPABASE_URL or not SUPABASE_KEY:
        return 0, "Supabase not configured"
    seen = {}
    for row in rows:
        ai = row.get('ai_code', '')
        if not ai: continue
        if ai not in seen:
            seen[ai] = row
        else:
            existing_score = sum(1 for v in seen[ai].values() if v and str(v) not in ('0',''))
            new_score = sum(1 for v in row.values() if v and str(v) not in ('0',''))
            if new_score > existing_score:
                seen[ai] = row

    client_rows = []
    for ai, row in seen.items():
        dob_val = row.get('inv_dob', '') or None
        if dob_val == '': dob_val = None
        client_rows.append({
            'ai_code':   ai,
            'pan':       row.get('pan_no', '') or '',
            'full_name': row.get('inv_name', '') or '',
            'mobile':    row.get('mobile_no', '') or '',
            'email':     row.get('email', '') or '',
            'dob':       dob_val,
        })

    if not client_rows:
        return 0, None

    url = f"{SUPABASE_URL}/rest/v1/clients?on_conflict=ai_code"
    pushed = 0
    for start in range(0, len(client_rows), 500):
        batch = client_rows[start:start+500]
        req = urllib.request.Request(
            url, data=json.dumps(batch).encode(), method="POST",
            headers={"Content-Type": "application/json", "apikey": SUPABASE_KEY,
                     "Authorization": f"Bearer {SUPABASE_KEY}",
                     "Prefer": "resolution=merge-duplicates,return=minimal"})
        try:
            with urllib.request.urlopen(req, timeout=60) as _:
                pushed += len(batch)
        except urllib.error.HTTPError as e:
            return pushed, e.read().decode()
        except Exception as e:
            return pushed, str(e)
    return pushed, None

def fetch_nav_from_supabase(isin_list):
    if not SUPABASE_URL or not SUPABASE_KEY:
        return {}, None
    nav_map = {}
    try:
        for i in range(0, len(isin_list), 100):
            batch = isin_list[i:i+100]
            isin_filter = ','.join(batch)
            path = f'/rest/v1/CAMS_KARVY_Contact?select=ISIN_NO,nav_value,nav_date&ISIN_NO=in.({isin_filter})'
            result, err = supabase_get(path)
            if err: continue
            if result:
                for row in result:
                    isin = str(row.get('ISIN_NO', '')).strip().upper()
                    nav_value = float(row.get('nav_value', 0) or 0)
                    nav_date = str(row.get('nav_date', '')).strip()
                    if isin and nav_value:
                        nav_map[isin] = {'nav_value': nav_value, 'nav_date': nav_date}
        return nav_map, None
    except Exception as e:
        return nav_map, str(e)

# ── Excel column labels and widths ──
COL_LABELS = {
    "ai_code":"AI Code","Folio No":"Folio No","inv_name":"Investor Name",
    "city":"City","pincode":"Pincode","product":"Product","sch_name":"Scheme Name",
    "rep_date":"Report Date","unit_balance":"Unit Balance",
    "total_amount_value":"Total Amount Value","jnt_name1":"Joint 1","jnt_name2":"Joint 2",
    "phone_off":"Phone Off","phone_res":"Phone Res","email":"Email",
    "holding_nature":"Holding","uin_no":"UIN","pan_no":"PAN",
    "joint1_pan":"Joint PAN 1","joint2_pan":"Joint PAN 2","guard_pan":"Guard PAN",
    "tax_status":"Tax Status","broker_code":"Broker","subbroker":"Sub-Broker",
    "reinv_flag":"Reinvest","bank_name":"Bank","branch":"Branch",
    "ac_type":"AC Type","ac_no":"AC Number","b_city":"Bank City",
    "b_pincode":"Bank Pincode","inv_dob":"DOB","mobile_no":"Mobile",
    "occupation":"Occupation","inv_iin":"IIN","nom_name":"Nominee 1",
    "relation":"Relation","nom_city":"Nom City","nom_state":"Nom State",
    "nom_pincode":"Nom Pincode","nom_ph_off":"Nom Ph Off","nom_ph_res":"Nom Ph Res",
    "nom_email":"Nom Email","nom_percentage":"Nom %","nom2_name":"Nominee 2",
    "nom2_relation":"Nom2 Rel","nom2_city":"Nom2 City","nom2_state":"Nom2 State",
    "nom2_pincode":"Nom2 Pincode","nom2_ph_off":"Nom2 Ph Off","nom2_ph_res":"Nom2 Ph Res",
    "nom2_email":"Nom2 Email","nom2_percentage":"Nom2 %","nom3_name":"Nominee 3",
    "nom3_relation":"Nom3 Rel","nom3_city":"Nom3 City","nom3_state":"Nom3 State",
    "nom3_pincode":"Nom3 Pincode","nom3_ph_off":"Nom3 Ph Off","nom3_ph_res":"Nom3 Ph Res",
    "nom3_email":"Nom3 Email","nom3_percentage":"Nom3 %","ifsc_code":"IFSC",
    "dp_id":"DP ID","demat":"Demat","guard_name":"Guardian","brokcode":"Brok Code",
    "folio_date":"Folio Date","aadhaar":"Aadhaar","tpa_linked":"TPA",
    "fh_ckyc_no":"FH CKYC","jh1_ckyc":"JH1 CKYC","jh2_ckyc":"JH2 CKYC",
    "g_ckyc_no":"G CKYC","jh1_dob":"JH1 DOB","jh2_dob":"JH2 DOB",
    "guardian_dob":"Guard DOB","amc_code":"AMC","gst_state_code":"GST State",
    "folio_old":"Old Folio","ISIN_NO":"ISIN No","country":"Country",
    "remarks":"Remarks","address":"Address","bank_address":"Bank Address",
    "nom_address":"Nom Address","ac_no_flag":"AC Flag","nav_value":"NAV Value",
    "nav_date":"NAV Date","Data_From":"Data From",
}
COL_WIDTHS = {
    'ai_code':9,'Folio No':15,'inv_name':28,'city':14,'pincode':9,
    'product':10,'sch_name':42,'rep_date':13,'unit_balance':14,'total_amount_value':20,
    'pan_no':14,'email':30,'mobile_no':14,'bank_name':24,'ac_no':20,
    'ifsc_code':14,'address':42,'bank_address':30,'nom_name':22,'ac_no_flag':14,
    'nav_value':12,'nav_date':14,
}

def _build_excel_buf(rows, stats):
    wb = Workbook()
    ws = wb.active
    ws.title = "Client Master"
    HDR_BG = "13171F"; HDR_FG = "00E5A0"; FLAG_C = "FF4455"
    for ci, col in enumerate(CLIENT_MASTER_COLUMNS, 1):
        c = ws.cell(row=1, column=ci, value=COL_LABELS.get(col, col.upper()))
        c.font = Font(name='Arial', bold=True, color=HDR_FG, size=9)
        c.fill = PatternFill("solid", start_color=HDR_BG)
        c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = "C2"
    for ri, row in enumerate(rows, 2):
        flag = row.get('ac_no_flag', '')
        for ci, col in enumerate(CLIENT_MASTER_COLUMNS, 1):
            v = row.get(col, '')
            if col in NUMERIC_COLS:
                try: v = float(v) if v else 0.0
                except: v = 0.0
            c = ws.cell(row=ri, column=ci, value=v if v != '' else None)
            if col == 'ac_no_flag' and flag:
                c.font = Font(name='Arial', bold=True, color=FLAG_C, size=9)
            else:
                c.font = Font(name='Arial', size=9)
    for ci, col in enumerate(CLIENT_MASTER_COLUMNS, 1):
        ws.column_dimensions[get_column_letter(ci)].width = COL_WIDTHS.get(col, 11)
    ws.auto_filter.ref = f"A1:{get_column_letter(len(CLIENT_MASTER_COLUMNS))}1"
    buf = io.BytesIO()
    wb.save(buf); buf.seek(0)
    return buf

# ── CAMS parser ──
def parse_cams_row(line):
    line = line.strip(); tokens = []; i = 0
    while i < len(line):
        if line[i] == "'":
            i += 1; val = []
            while i < len(line):
                if line[i] == "'":
                    if i+1 >= len(line) or line[i+1] == ',': i += 1; break
                    else: val.append(line[i]); i += 1
                else: val.append(line[i]); i += 1
            tokens.append(''.join(val).strip())
            if i < len(line) and line[i] == ',': i += 1
        elif line[i] == ',': tokens.append(''); i += 1
        else:
            j = i
            while j < len(line) and line[j] != ',' and line[j] != "'": j += 1
            tokens.append(line[i:j].strip()); i = j
            if i < len(line) and line[i] == ',': i += 1
    return tokens

def read_cams_file(raw_bytes):
    lines = [p.decode('utf-8-sig', errors='replace') for p in raw_bytes.split(b'\x0d\x0a') if p.strip()]
    if not lines: return [], []
    header = [h.upper().strip() for h in parse_cams_row(lines[0])]
    rows = []
    for line in lines[1:]:
        if not line.strip(): continue
        vals = parse_cams_row(line)
        while len(vals) < len(header): vals.append('')
        rows.append({header[i]: vals[i] for i in range(len(header))})
    return header, rows

def detect_cams(raw_bytes):
    first = raw_bytes.split(b'\x0d\x0a')[0].decode('utf-8-sig', errors='replace')
    return "'FOLIOCHK'" in first or "'INV_NAME'" in first

def process_cams_bytes(raw_bytes):
    if not detect_cams(raw_bytes):
        return None, None, "Not a CAMS file"
    _, cams_rows = read_cams_file(raw_bytes)
    if not cams_rows:
        return None, None, "No data rows found"

    pan_to_ai, name_dob_to_ai, folio_to_ai, max_ai = fetch_existing_pan_map()
    ai_counter = max_ai + 1
    sup_count  = len(pan_to_ai)

    unique_isins = set()
    for row in cams_rows:
        prodcode = g(row, 'PRODUCT').strip().upper()
        isin = prodcode_to_isin(prodcode)
        if isin: unique_isins.add(isin)

    nav_map = {}
    nav_warnings = []
    if unique_isins:
        nav_map, nav_err = fetch_nav_from_supabase(list(unique_isins))
        if nav_err: nav_warnings.append(f"NAV warning: {nav_err}")

    mapped, errors = [], []
    for i, row in enumerate(cams_rows):
        try:
            pan      = g(row,'PAN_NO').upper()
            inv_name = g(row,'INV_NAME')
            inv_dob  = clean_date(g(row,'INV_DOB'))
            raw_ac   = g(row,'AC_NO')
            ac_no    = clean_ac_no(raw_ac)
            folio    = clean_folio(g(row,'FOLIOCHK'))
            product  = clean_val(g(row,'PRODUCT'))

            ai_code, pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter = get_or_create_ai(
                pan, inv_name, inv_dob, folio,
                pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter)

            holding = HOLDING_MAP.get(g(row,'HOLDING_NATURE').upper(), g(row,'HOLDING_NATURE'))
            tax     = TAX_STATUS_MAP.get(g(row,'TAX_STATUS').upper(), g(row,'TAX_STATUS'))
            ac_type = AC_TYPE_MAP.get(g(row,'AC_TYPE').upper(), g(row,'AC_TYPE'))
            occ     = OCC_MAP.get(g(row,'OCCUPATION'), g(row,'OCCUPATION'))

            prodcode  = product.strip().upper()
            isin      = prodcode_to_isin(prodcode)
            nav_info  = nav_map.get(isin, {}) if isin else {}
            nav_value = nav_info.get('nav_value', 0.0)
            nav_date  = nav_info.get('nav_date', '')
            unit_bal  = _sf(g(row,'CLOS_BAL') or '0')
            total_amt = round(nav_value * unit_bal, 4) if nav_value and unit_bal else 0.0

            if not isin and product:     nav_flag = 'PRODCODE UNMAPPED'
            elif isin and not nav_value: nav_flag = 'NAV MISSING'
            else:                        nav_flag = ''

            r = {col:'' for col in CLIENT_MASTER_COLUMNS}
            r.update({
                'ai_code':ai_code,'Folio No':folio,'inv_name':inv_name,
                'city':g(row,'CITY'),'pincode':g(row,'PINCODE'),
                'product':product,'sch_name':g(row,'SCH_NAME'),
                'rep_date':clean_date(g(row,'REP_DATE')),
                'unit_balance':str(unit_bal),'total_amount_value':str(total_amt),
                'nav_value':str(nav_value),'nav_date':nav_date,
                'jnt_name1':g(row,'JNT_NAME1'),'jnt_name2':g(row,'JNT_NAME2'),
                'phone_off':clean_phone(g(row,'PHONE_OFF')),
                'phone_res':clean_phone(g(row,'PHONE_RES')),
                'email':g(row,'EMAIL').lower(),'holding_nature':holding,
                'uin_no':g(row,'UIN_NO'),'pan_no':pan,
                'joint1_pan':g(row,'JOINT1_PAN').upper(),
                'joint2_pan':g(row,'JOINT2_PAN').upper(),
                'guard_pan':g(row,'GUARD_PAN').upper(),
                'tax_status':tax,'broker_code':g(row,'BROKER_CODE'),
                'subbroker':g(row,'SUBBROKER'),'reinv_flag':g(row,'REINV_FLAG'),
                'bank_name':g(row,'BANK_NAME'),'branch':g(row,'BRANCH'),
                'ac_type':ac_type,'ac_no':ac_no,'b_city':g(row,'B_CITY'),
                'b_pincode':g(row,'B_PINCODE'),
                'inv_dob':inv_dob,'mobile_no':clean_phone(g(row,'MOBILE_NO')),
                'occupation':occ,'inv_iin':g(row,'INV_IIN') or '0',
                'nom_name':g(row,'NOM_NAME'),'relation':g(row,'RELATION'),
                'nom_city':g(row,'NOM_CITY'),'nom_state':g(row,'NOM_STATE'),
                'nom_pincode':g(row,'NOM_PINCODE'),
                'nom_ph_off':clean_phone(g(row,'NOM_PH_OFF')),
                'nom_ph_res':clean_phone(g(row,'NOM_PH_RES')),
                'nom_email':g(row,'NOM_EMAIL').lower(),
                'nom_percentage':g(row,'NOM_PERCENTAGE') or '0',
                'nom2_name':g(row,'NOM2_NAME'),'nom2_relation':g(row,'NOM2_RELATION'),
                'nom2_city':g(row,'NOM2_CITY'),'nom2_state':g(row,'NOM2_STATE'),
                'nom2_pincode':g(row,'NOM2_PINCODE'),
                'nom2_ph_off':clean_phone(g(row,'NOM2_PH_OFF')),
                'nom2_ph_res':clean_phone(g(row,'NOM2_PH_RES')),
                'nom2_email':g(row,'NOM2_EMAIL').lower(),
                'nom2_percentage':g(row,'NOM2_PERCENTAGE') or '0',
                'nom3_name':g(row,'NOM3_NAME'),'nom3_relation':g(row,'NOM3_RELATION'),
                'nom3_city':g(row,'NOM3_CITY'),'nom3_state':g(row,'NOM3_STATE'),
                'nom3_pincode':g(row,'NOM3_PINCODE'),
                'nom3_ph_off':clean_phone(g(row,'NOM3_PH_OFF')),
                'nom3_ph_res':clean_phone(g(row,'NOM3_PH_RES')),
                'nom3_email':g(row,'NOM3_EMAIL').lower(),
                'nom3_percentage':g(row,'NOM3_PERCENTAGE') or '0',
                'ifsc_code':clean_ifsc(g(row,'IFSC_CODE')),
                'dp_id':g(row,'DP_ID'),'demat':g(row,'DEMAT') or 'N',
                'guard_name':g(row,'GUARD_NAME'),
                'brokcode':g(row,'BROKCODE') or g(row,'BROKER_CODE'),
                'folio_date':clean_date(g(row,'FOLIO_DATE')),
                'aadhaar':g(row,'AADHAAR'),'tpa_linked':g(row,'TPA_LINKED'),
                'fh_ckyc_no':g(row,'FH_CKYC_NO'),'jh1_ckyc':g(row,'JH1_CKYC'),
                'jh2_ckyc':g(row,'JH2_CKYC'),'g_ckyc_no':g(row,'G_CKYC_NO'),
                'jh1_dob':clean_date(g(row,'JH1_DOB')),
                'jh2_dob':clean_date(g(row,'JH2_DOB')),
                'guardian_dob':clean_date(g(row,'GUARDIAN_DOB')),
                'amc_code':g(row,'AMC_CODE'),'gst_state_code':g(row,'GST_STATE_CODE'),
                'folio_old':g(row,'FOLIO_OLD'),'ISIN_NO':isin,
                'country':g(row,'COUNTRY') or 'India','remarks':g(row,'REMARKS'),
                'address':combine_addr(g(row,'ADDRESS1'),g(row,'ADDRESS2'),g(row,'ADDRESS3')),
                'bank_address':combine_addr(g(row,'B_ADDRESS1'),g(row,'B_ADDRESS2'),g(row,'B_ADDRESS3')),
                'nom_address':combine_addr(g(row,'NOM_ADDR1'),g(row,'NOM_ADDR2'),g(row,'NOM_ADDR3')),
                'ac_no_flag':fix_ac_flag(raw_ac, ac_no) or nav_flag,
                'Data_From': 'CAMS',
            })
            mapped.append(r)
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")

    mapped, removed = dedup_rows(mapped)
    unique = len(set(r.get('ai_code','') for r in mapped))
    new_cl = sum(1 for r in mapped if ai_num(r.get('ai_code','')) > max_ai)
    stats  = {
        'total_rows': len(mapped), 'removed_duplicates': removed,
        'unique_clients': unique, 'existing_clients': unique - new_cl,
        'new_clients': new_cl, 'supabase_pan_count': sup_count,
        'with_pan':     sum(1 for r in mapped if PAN_PATTERN.match(r.get('pan_no','').upper())),
        'with_balance': sum(1 for r in mapped if _sf(r.get('total_amount_value','0')) > 0),
        'flagged_ac':   sum(1 for r in mapped if r.get('ac_no_flag','')),
        'nav_matched':  sum(1 for r in mapped if _sf(r.get('nav_value','0')) > 0),
        'errors': errors[:20] + nav_warnings,
    }
    return mapped, stats, None

def process_karvy(new_karvy_bytes, karvy_master_bytes):
    try:
        nk = pd.read_csv(io.BytesIO(new_karvy_bytes), dtype=str).fillna('')
        km = pd.read_csv(io.BytesIO(karvy_master_bytes), dtype=str).fillna('')
    except Exception as e:
        return None, None, f"Could not read CSV: {e}"

    if 'Folio Number' not in nk.columns:
        return None, None, "File 1 missing 'Folio Number'"
    if 'Folio' not in km.columns or 'BankAccno' not in km.columns:
        return None, None, "File 2 missing 'Folio' or 'BankAccno'"

    nk['Folio Number'] = nk['Folio Number'].str.strip()
    km['Folio']        = km['Folio'].str.strip()

    bank_cols_wanted = ['Folio','BankAccno','Bank Name','Account Type','Branch',
        'Bank Address #1','Bank Address #2','Bank Address #3','Bank City',
        'Date of Birth','Tax Status','Occ Code','Mode of Holding',
        'Mapin Id','Holder 1 Aadhaar info','PAN Number','Mobile Number']
    bank_cols = [c for c in bank_cols_wanted if c in km.columns]
    km_bank = km[bank_cols].drop_duplicates(subset='Folio', keep='first')
    km_bank = km_bank.rename(columns={
        'Bank Name':'bk_bank_name','Account Type':'bk_ac_type','Branch':'bk_branch',
        'Bank Address #1':'bk_addr1','Bank Address #2':'bk_addr2','Bank Address #3':'bk_addr3',
        'Bank City':'bk_city','Date of Birth':'bk_dob','Tax Status':'bk_tax_status',
        'Occ Code':'bk_occ_code','Mode of Holding':'bk_holding','Mapin Id':'bk_mapin',
        'Holder 1 Aadhaar info':'bk_aadhaar','PAN Number':'bk_pan','Mobile Number':'bk_mobile',
    })
    merged = nk.merge(km_bank, left_on='Folio Number', right_on='Folio', how='left').fillna('')

    pan_to_ai, name_dob_to_ai, folio_to_ai, max_ai = fetch_existing_pan_map()
    ai_counter = max_ai + 1

    unique_isins = set()
    for _, row in merged.iterrows():
        isin_direct = _s(row.get('SchemeISIN', '')).strip().upper()
        if isin_direct: unique_isins.add(isin_direct)
        else:
            prodcode = _s(row.get('Product Code', '')).strip().upper()
            isin = prodcode_to_isin(prodcode)
            if isin: unique_isins.add(isin)

    nav_map, nav_warnings = {}, []
    if unique_isins:
        nav_map, nav_err = fetch_nav_from_supabase(list(unique_isins))
        if nav_err: nav_warnings.append(f"NAV warning: {nav_err}")

    mapped, errors = [], []
    for i, row in merged.iterrows():
        try:
            pan      = (_s(row.get('PAN','')) or _s(row.get('bk_pan',''))).upper()
            inv_name = _s(row.get('Investor Name','')).upper()
            inv_dob  = clean_date(_s(row.get('bk_dob','')))
            raw_ac   = _s(row.get('BankAccno',''))
            ac_no    = clean_ac_no(raw_ac)
            folio    = clean_folio(_s(row.get('Folio Number','')))
            product  = _s(row.get('Product Code','')).strip()
            broker   = _s(row.get('Broker Code','')) or _s(row.get('Agent Code',''))

            ai_code, pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter = get_or_create_ai(
                pan, inv_name, inv_dob, folio,
                pan_to_ai, name_dob_to_ai, folio_to_ai, ai_counter)

            holding = KARVY_HOLDING_MAP.get(_s(row.get('bk_holding','')), '')
            tax     = TAX_STATUS_MAP.get(_s(row.get('bk_tax_status','')).upper(), _s(row.get('bk_tax_status','')))
            ac_type = KARVY_AC_TYPE_MAP.get(_s(row.get('bk_ac_type','')).upper(), '')
            occ     = KARVY_OCC_MAP.get(_s(row.get('bk_occ_code','')), _s(row.get('Occupation Description','')))

            isin_direct = _s(row.get('SchemeISIN', '')).strip().upper()
            isin = isin_direct if isin_direct else prodcode_to_isin(product.strip().upper())

            nav_info  = nav_map.get(isin, {}) if isin else {}
            nav_value = nav_info.get('nav_value', 0.0)
            nav_date  = nav_info.get('nav_date', '')
            unit_bal  = _sf(_clean_num(_s(row.get('Balance',''))))
            total_amt = round(nav_value * unit_bal, 4) if nav_value and unit_bal else 0.0

            if not isin and product:     nav_flag = 'PRODCODE UNMAPPED'
            elif isin and not nav_value: nav_flag = 'NAV MISSING'
            else:                        nav_flag = ''

            mobile = clean_phone(_s(row.get('Mobile No','')) or _s(row.get('bk_mobile','')))

            r = {col: '' for col in CLIENT_MASTER_COLUMNS}
            r.update({
                'ai_code':ai_code,'Folio No':folio,'inv_name':inv_name,
                'city':_s(row.get('City','')),'pincode':_s(row.get('Pincode','')),
                'product':product,'sch_name':_s(row.get('Fund Description','')),
                'rep_date':clean_date(_s(row.get('Report Date',''))),
                'unit_balance':str(unit_bal),'total_amount_value':str(total_amt),
                'nav_value':str(nav_value),'nav_date':nav_date,
                'jnt_name1':_s(row.get('Joint Name 1','')),'jnt_name2':_s(row.get('Joint Name 2','')),
                'phone_off':clean_phone(_s(row.get('Phone Office',''))),
                'phone_res':clean_phone(_s(row.get('Phone Residence',''))),
                'email':_s(row.get('Email','')).lower(),'holding_nature':holding,
                'uin_no':_s(row.get('bk_mapin','')),'pan_no':pan,
                'tax_status':tax,'broker_code':broker,'brokcode':broker,
                'bank_name':_s(row.get('bk_bank_name','')),'branch':_s(row.get('bk_branch','')),
                'ac_type':ac_type,'ac_no':ac_no,'b_city':_s(row.get('bk_city','')),
                'inv_dob':inv_dob,'mobile_no':mobile,
                'occupation':occ,'inv_iin':_s(row.get('Investor ID','')) or '0',
                'nom_percentage':'0','nom2_percentage':'0','nom3_percentage':'0',
                'ifsc_code':'','dp_id':_s(row.get('DPID','')),'demat':'N',
                'aadhaar':_s(row.get('bk_aadhaar','')),'country':_s(row.get('Country','')) or 'India',
                'amc_code':_s(row.get('Fund','')),'ISIN_NO':isin,
                'address':_combine(row.get('Address #1',''),row.get('Address #2',''),row.get('Address #3','')),
                'bank_address':_combine(row.get('bk_addr1',''),row.get('bk_addr2',''),row.get('bk_addr3','')),
                'ac_no_flag':fix_ac_flag(raw_ac, ac_no) or nav_flag,
                'Data_From':'KARVY',
            })
            mapped.append(r)
        except Exception as e:
            errors.append(f"Row {i+2}: {e}")

    mapped, removed = dedup_rows(mapped)
    unique = len(set(r.get('ai_code','') for r in mapped))
    new_cl = sum(1 for r in mapped if ai_num(r.get('ai_code','')) > max_ai)
    stats = {
        'total_rows':len(mapped),'removed_duplicates':removed,
        'unique_clients':unique,'existing_clients':unique-new_cl,'new_clients':new_cl,
        'with_pan':     sum(1 for r in mapped if PAN_PATTERN.match(r.get('pan_no','').upper())),
        'with_balance': sum(1 for r in mapped if _sf(r.get('total_amount_value','0')) > 0),
        'flagged_ac':   sum(1 for r in mapped if r.get('ac_no_flag','')),
        'nav_matched':  sum(1 for r in mapped if _sf(r.get('nav_value','0')) > 0),
        'errors': errors[:20] + nav_warnings,
    }
    return mapped, stats, None

def _push_rows_to_supabase(rows, session_key):
    """Common push logic for both CAMS and KARVY"""
    if not SUPABASE_URL or not SUPABASE_KEY:
        return jsonify({'error': 'Supabase not configured'}), 400

    sd = session_data.get('default', {})
    if session_key not in sd:
        return jsonify({'error': 'No Excel data found. Upload the reviewed Excel first.'}), 400

    rows = sd[session_key]
    if not rows: return jsonify({'error': 'No rows to push'}), 400

    seen = {}
    for row in rows:
        key = (row.get('ai_code',''), row.get('Folio No',''), row.get('product',''))
        if key not in seen: seen[key] = row
        elif float(row.get('total_amount_value','0') or 0) > \
             float(seen[key].get('total_amount_value','0') or 0):
            seen[key] = row
    rows = list(seen.values())

    synced, sync_err = sync_clients_table(rows)
    if sync_err:
        return jsonify({'error': f'clients sync failed: {sync_err}'}), 500

    pushed, BATCH = 0, 500
    url = f"{SUPABASE_URL}/rest/v1/CAMS_KARVY_Contact?on_conflict=ai_code,Folio%20No,product"
    for start in range(0, len(rows), BATCH):
        batch = rows[start:start+BATCH]
        req = urllib.request.Request(url, data=json.dumps(batch).encode(), method="POST",
            headers={"Content-Type":"application/json","apikey":SUPABASE_KEY,
                     "Authorization":f"Bearer {SUPABASE_KEY}",
                     "Prefer":"resolution=merge-duplicates,return=minimal"})
        try:
            with urllib.request.urlopen(req, timeout=60) as _: pushed += len(batch)
        except urllib.error.HTTPError as e:
            return jsonify({'error': f'Batch {start//BATCH+1}: {e.read().decode()}'}), 500
        except Exception as e:
            return jsonify({'error': str(e)}), 500

    return jsonify({'success': True, 'message': f'✅ {pushed} rows pushed, {synced} clients synced'})

def _parse_excel_upload(file):
    """Parse uploaded Excel file, return (rows, error)"""
    wb = load_workbook(io.BytesIO(file.read()), data_only=True)
    ws = wb.active
    label_to_key = {v.lower(): k for k, v in COL_LABELS.items()}
    raw_headers  = [str(cell.value).strip().lower() if cell.value else '' for cell in ws[1]]
    col_map = []
    for ci, h in enumerate(raw_headers):
        if h in CLIENT_MASTER_COLUMNS: col_map.append((ci, h))
        elif h in label_to_key:        col_map.append((ci, label_to_key[h]))
    if not col_map:
        return None, 'Could not match any columns. Upload the Excel exported from this tool.'
    rows = []
    for row_vals in ws.iter_rows(min_row=2, values_only=True):
        if not any(row_vals): continue
        r = {col: '' for col in CLIENT_MASTER_COLUMNS}
        for ci, key in col_map:
            if ci >= len(row_vals): continue
            v = row_vals[ci]
            if v is None: r[key] = ''; continue
            if key in NUMERIC_COLS:
                try:
                    fv = float(v); r[key] = str(int(fv)) if fv == int(fv) else str(round(fv, 4))
                except: r[key] = '0'
            else:
                sv = str(v).strip()
                if sv in ('None','nan','NaN',''): r[key] = ''; continue
                if sv.endswith('.0') and sv[:-2].isdigit(): sv = sv[:-2]
                r[key] = sv
        rows.append(r)
    if not rows: return None, 'No data rows found in Excel'
    return rows, None

# ══════════════════════════════════════════════
# UPLOAD TOOL API ROUTES
# ══════════════════════════════════════════════

# ── CAMS ──
@app.route('/upload/process', methods=['POST'])
def upload_process():
    if 'file' not in request.files:
        return jsonify({'error': 'No CAMS file uploaded'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Please upload a .csv file'}), 400
    mapped, stats, err = process_cams_bytes(f.read())
    if err: return jsonify({'error': err}), 400
    excel_buf = _build_excel_buf(mapped, stats)
    session_data.setdefault('default', {})['cams_data'] = {
        'rows': mapped, 'stats': stats, 'excel_bytes': excel_buf.getvalue(),
    }
    return jsonify({**stats, 'ready': True})

@app.route('/upload/download-excel')
def upload_download_excel():
    sd = session_data.get('default', {})
    if 'cams_data' not in sd:
        return "No data ready. Please process a CAMS CSV file first.", 400
    data = sd['cams_data']
    buf = io.BytesIO(data['excel_bytes']) if 'excel_bytes' in data \
          else _build_excel_buf(data['rows'], data['stats'])
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='client_master_review.xlsx')

@app.route('/upload/preview-excel', methods=['POST'])
def upload_preview_excel():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Please upload the .xlsx file'}), 400
    rows, err = _parse_excel_upload(f)
    if err: return jsonify({'error': err}), 400
    session_data.setdefault('default', {})['cams_excel'] = rows
    prev_cols = ['ai_code','Folio No','inv_name','pan_no','product',
                 'sch_name','rep_date','unit_balance','total_amount_value','nav_value','nav_date','ac_no_flag']
    preview = [{c: r.get(c,'') for c in prev_cols} for r in rows[:10]]
    return jsonify({'total_rows': len(rows),
                    'flagged_ac': sum(1 for r in rows if r.get('ac_no_flag','')),
                    'preview': preview, 'preview_cols': prev_cols, 'ready': True})

@app.route('/upload/push', methods=['POST'])
def upload_push():
    return _push_rows_to_supabase(session_data.get('default',{}).get('cams_excel',[]), 'cams_excel')

# ── KARVY ──
@app.route('/upload/karvy/process', methods=['POST'])
def upload_karvy_process():
    if 'file_new' not in request.files or 'file_master' not in request.files:
        return jsonify({'error': 'Upload both files'}), 400
    f1, f2 = request.files['file_new'], request.files['file_master']
    if not f1.filename.lower().endswith('.csv') or not f2.filename.lower().endswith('.csv'):
        return jsonify({'error': 'Both files must be .csv'}), 400
    mapped, stats, err = process_karvy(f1.read(), f2.read())
    if err: return jsonify({'error': err}), 400
    excel_buf = _build_excel_buf(mapped, stats)
    session_data.setdefault('default', {})['karvy_data'] = {
        'rows': mapped, 'stats': stats, 'excel_bytes': excel_buf.getvalue(),
    }
    return jsonify({**stats, 'ready': True})

@app.route('/upload/karvy/download-excel')
def upload_karvy_download():
    sd = session_data.get('default', {})
    if 'karvy_data' not in sd:
        return "No KARVY data. Process KARVY files first.", 400
    data = sd['karvy_data']
    buf = io.BytesIO(data['excel_bytes']) if 'excel_bytes' in data \
          else _build_excel_buf(data['rows'], data['stats'])
    return send_file(buf,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True, download_name='karvy_review.xlsx')

@app.route('/upload/karvy/preview-excel', methods=['POST'])
def upload_karvy_preview():
    if 'file' not in request.files:
        return jsonify({'error': 'No file uploaded'}), 400
    f = request.files['file']
    if not f.filename.lower().endswith('.xlsx'):
        return jsonify({'error': 'Please upload the .xlsx file'}), 400
    rows, err = _parse_excel_upload(f)
    if err: return jsonify({'error': err}), 400
    session_data.setdefault('default', {})['karvy_excel'] = rows
    prev_cols = ['ai_code','Folio No','inv_name','pan_no','product',
                 'sch_name','rep_date','unit_balance','total_amount_value','nav_value','nav_date','ac_no_flag']
    preview = [{c: r.get(c,'') for c in prev_cols} for r in rows[:10]]
    return jsonify({'total_rows': len(rows),
                    'flagged_ac': sum(1 for r in rows if r.get('ac_no_flag','')),
                    'preview': preview, 'preview_cols': prev_cols, 'ready': True})

@app.route('/upload/karvy/push', methods=['POST'])
def upload_karvy_push():
    return _push_rows_to_supabase(session_data.get('default',{}).get('karvy_excel',[]), 'karvy_excel')

# ══════════════════════════════════════════════
# CATCH-ALL — serve website static assets
# (CSS, JS, images from assets/ folder)
# ══════════════════════════════════════════════
@app.route('/<path:path>')
def static_files(path):
    if os.path.isfile(path):
        return send_from_directory('.', path)
    return send_from_directory('.', 'index.html')

# ══════════════════════════════════════════════
# ENTRY POINT
# ══════════════════════════════════════════════
if __name__ == '__main__':
    port = int(os.environ.get('PORT', 5000))
    app.run(host='0.0.0.0', port=port, debug=False)
